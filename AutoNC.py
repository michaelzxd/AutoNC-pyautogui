import os
import pyautogui
import openpyxl
import time
import pyperclip
pyautogui.PAUSE= 0.6
pyautogui.FAILSAFE = True

wb = openpyxl.load_workbook('LEDGER.xlsx',read_only=True, data_only=True)
NCsheet = wb['NC']

payment_type_dict = {'正价采购': '0101', '估价采购': '0201','冲估价采购': '0202','估价转正价采购': '0102','估价转估价采购': '0203'}
operator_name = str(NCsheet.cell(row=7,column=3).value)

def English_check():
	"step 1: check input status"
	check_input_method = pyautogui.confirm(text='请确认是否已切换至英文小写输入法', buttons=['已切换', '切换后再回来'])
	if check_input_method == '已切换':
		pass
	else:
		exit()

"my paste definition"
def my_paste(text):
    pyperclip.copy(text)
    pyautogui.hotkey('ctrl','v')

"mouse move definition I am most proud of"
def locate_move_and_click(target_image,x_move,y_move):
	try:
		pyautogui.screenshot()
		(x,y)=pyautogui.center(pyautogui.locateOnScreen(target_image,confidence=0.8))
		pyautogui.moveTo(x, y, 0.5, pyautogui.easeOutQuad) 
		pyautogui.move(x_move, y_move)
		pyautogui.click()
	except:
		time.sleep(2)
		pyautogui.screenshot()
		(x,y)=pyautogui.center(pyautogui.locateOnScreen(target_image,confidence=0.8))
		pyautogui.moveTo(x, y, 0.5, pyautogui.easeOutQuad) 
		pyautogui.move(x_move, y_move)
		pyautogui.click()



def go_to_contract_page_from_start():
	locate_move_and_click("contract_button.png",0,0)
	time.sleep(3)
	pyautogui.hotkey('ctrl', 'n')
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('enter')
	time.sleep(2)
	locate_move_and_click("qg_position.png",300,0)
	

def find_contract_page(f_QCnumber):
	time.sleep(1)
	pyautogui.typewrite(f_QCnumber, interval=0.15)
	pyautogui.press('enter')
	time.sleep(1)
	pyautogui.hotkey('alt', 'y')
	time.sleep(2)
	pyautogui.press('tab')
	pyautogui.press('space')
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'e')
	time.sleep(2)

def fulfilling_contract_page(f_dealdate,f_operator_name,f_their_contract_number,f_our_contract_number,f_paymentrange):
	time.sleep(1)
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.typewrite(f_dealdate, interval=0.15)
	pyautogui.press('enter')
	locate_move_and_click("operator.png",160, 0)
	my_paste(f_operator_name)
	pyautogui.press('enter')
	pyautogui.typewrite(f_their_contract_number, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_our_contract_number, interval=0.15)
	pyautogui.press('enter')
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('enter')
	pyautogui.typewrite("001", interval=0.15) 
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.typewrite(f_paymentrange, interval=0.15)
	time.sleep(1)
	pyautogui.press('down')
	pyautogui.press('enter')
	pyautogui.hotkey('ctrl', 's')
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'u')

def go_to_next_page(get_number,next_button):
	time.sleep(2)
	pyautogui.screenshot()
	(x,y)=pyautogui.center(pyautogui.locateOnScreen(get_number,confidence=0.8))
	pyautogui.moveTo(x, y, 0.5, pyautogui.easeOutQuad)
	pyautogui.move(130, 0)
	pyautogui.doubleClick()
	pyautogui.hotkey('ctrl', 'c')
	time.sleep(1)
	locate_move_and_click("procedures.png",0,0)
	time.sleep(1)
	locate_move_and_click(next_button,0,0)
	time.sleep(3)

def find_dh_page():
	locate_move_and_click("add_page.png",50,22)
	time.sleep(2)
	locate_move_and_click("lianyue.png",0,0)
	pyautogui.press('down')
	pyautogui.press('backspace')
	pyautogui.press(['backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace'])
	pyautogui.press('down')
	pyautogui.press('space')
	pyautogui.hotkey('ctrl', 'v')
	pyautogui.press('enter')
	pyautogui.hotkey('alt', 'y')
	time.sleep(3)
	locate_move_and_click("enter_choose.png",0,0)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'e')
	time.sleep(3)

def fulfilling_dh_page(f_dhdate,f_barge_name,f_tank_name,f_quantity_tons,f_quantity_bls):
	pyautogui.typewrite(f_dhdate, interval=0.15)
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.press('backspace')
	pyautogui.press(['backspace','backspace','backspace'])
	my_paste(f_barge_name)
	pyautogui.press('down')
	pyautogui.press('enter')
	pyautogui.press('backspace')
	pyautogui.press(['backspace','backspace','backspace'])
	my_paste(f_tank_name)
	pyautogui.press('down')
	pyautogui.press('enter')
	locate_move_and_click("ratio_position.png",0,20)
	pyautogui.press(['enter','enter','enter','enter','enter','left','left','left','left','left'])
	pyautogui.typewrite(f_quantity_tons, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_tons, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_bls, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_bls, interval=0.15)
	pyautogui.press('enter')
	pyautogui.hotkey('ctrl', 's')
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'u')


def go_back_to_DH():
	time.sleep(1)
	locate_move_and_click("procedures.png",0,0)
	time.sleep(1)
	locate_move_and_click("dh_button.png",0,0)
	time.sleep(3)
	locate_move_and_click("add_page.png",50,22)
	time.sleep(2)
	pyautogui.hotkey('alt', 'y')
	time.sleep(2)
	locate_move_and_click("enter_choose.png",0,0)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'e')
	time.sleep(2)

def close_pages():
	locate_move_and_click('close_cr.png',33,0)
	locate_move_and_click('close_dh.png',41,0)


def iterate_find_cr_page():
	time.sleep(1)
	pyautogui.hotkey('alt', '6')
	time.sleep(3)
	locate_move_and_click('enter_position.png',0,20)
	pyautogui.doubleClick()
	pyautogui.press('backspace')
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('space')
	pyautogui.hotkey('ctrl', 'v')
	pyautogui.press('enter')
	pyautogui.hotkey('alt', 'y')
	time.sleep(2)
	locate_move_and_click("ite_cr_enter.png",0,0)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)

def iterate_fulfilling_cr_page(f_quantity_tons,f_quantity_bls,f_bldate,f_paymentdate):
	time.sleep(1)
	locate_move_and_click("tank_name.png",226,0)
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('enter')
	time.sleep(1)
	locate_move_and_click("ratio_position.png",0,23)
	time.sleep(1)
	pyautogui.doubleClick()
	time.sleep(1)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_tons, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_tons, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_quantity_bls, interval=0.15)
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.typewrite(f_bldate, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_bldate, interval=0.15)
	pyautogui.press('enter')
	pyautogui.press('enter')
	pyautogui.typewrite(f_paymentdate, interval=0.15)
	pyautogui.press('enter')
	pyautogui.hotkey('ctrl', 's')
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'g')
	time.sleep(1)


def CD_start_find_dh_page(f_CDnumber):
	time.sleep(1)
	locate_move_and_click("procedures.png",0,0)
	time.sleep(1)
	locate_move_and_click("dh_button.png",0,0)
	time.sleep(2)
	locate_move_and_click("add_page.png",50,22)
	time.sleep(2)
	locate_move_and_click("lianyue.png",0,0)
	pyautogui.press('down')
	pyautogui.press('backspace')
	pyautogui.press(['backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace'])
	pyautogui.press('down')
	pyautogui.press('space')
	pyautogui.typewrite(f_CDnumber, interval=0.2)
	pyautogui.press('enter')
	pyautogui.hotkey('alt', 'y')
	time.sleep(2)
	locate_move_and_click("enter_choose.png",0,0)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)
	pyautogui.hotkey('ctrl', 'e')
	time.sleep(3)



def first_lot_all_the_way():
	"step 1: wait to click on the procedure page"
	time.sleep(2)
	"step 2: go to the contract page"
	go_to_contract_page_from_start()
	"step 3: find the right contract page"
	find_contract_page(QCnumber)
	"step 4: start fulfilling contract page"
	fulfilling_contract_page(dealdate,operator_name,their_contract_number,our_contract_number,paymentrange)
	"step 5: go to dh page"
	go_to_next_page("contract_number.png","dh_button.png")
	"step 6: find the right DH page"
	find_dh_page()
	"step 7: start fulfilling dh page"
	fulfilling_dh_page(dhdate,barge_name,tank_name,quantity_tons,quantity_bls)
	"step 8: go to cr page"
	go_to_next_page("dh_number_position.png","cr_button.png")
	"step 9: find the right cr page"
	iterate_find_cr_page()
	"step 10: start fulfilling CR page"
	iterate_fulfilling_cr_page(quantity_tons,quantity_bls,bldate,paymentdate)



def first_iterate_from_DH_to_CR():
	"iterate all lots"
	for i in range(2,NCsheet.max_column+1):
		column_number = i + 2
		dhdate = str(NCsheet.cell(row=9,column=column_number).value)
		barge_name = str(NCsheet.cell(row=10,column=column_number).value)
		tank_name= str(NCsheet.cell(row=11,column=column_number).value)
		quantity_tons = str(NCsheet.cell(row=12,column=column_number).value)
		quantity_bls = str(NCsheet.cell(row=13,column=column_number).value)
		bldate = str(NCsheet.cell(row=15,column=column_number).value)
		paymentdate = str(NCsheet.cell(row=16,column=column_number).value)
		seller_invoice_number= str(NCsheet.cell(row=18,column=column_number).value)
		total_amount = str(NCsheet.cell(row=19,column=column_number).value)
		payment_date = str(NCsheet.cell(row=20,column=column_number).value)
		invoice_date = str(NCsheet.cell(row=21,column=column_number).value)
		payment_type = str(NCsheet.cell(row=22,column=column_number).value)

		if quantity_tons != "0":
			close_pages()
			"step 1: go back to DH to begin iterate"
			go_back_to_DH()
			"step 2: start fulfilling dh page"
			fulfilling_dh_page(dhdate,barge_name,tank_name,quantity_tons,quantity_bls)
			"step 3: go to cr page"
			go_to_next_page("dh_number_position.png","cr_button.png")
			"step 4: find the right cr page"
			iterate_find_cr_page()
			"step 5: start fulfilling CR page"
			iterate_fulfilling_cr_page(quantity_tons,quantity_bls,bldate,paymentdate)
			
		else:
			pyautogui.alert("亲，都搞定了哦，请在台账里保存CD和CR号码（必须的）")
			exit()


def new_iterate_from_DH_to_CR():
	if quantity_tons != "0":
		try:
			close_pages()
		except:
			pass
		"step 1: go to next DH to resume iterate"
		CD_start_find_dh_page(CDnumber)
		"step 2: start fulfilling dh page"
		fulfilling_dh_page(dhdate,barge_name,tank_name,quantity_tons,quantity_bls)
		"step 3: go to cr page"
		go_to_next_page("dh_number_position.png","cr_button.png")
		"step 4: find the right cr page"
		iterate_find_cr_page()
		"step 5: start fulfilling CR page"
		iterate_fulfilling_cr_page(quantity_tons,quantity_bls,bldate,paymentdate)
		time.sleep(2)
	else:
		pyautogui.alert("亲，都搞定了哦，请在台账里保存CR号码（必须的）")
		exit()

"invoice and print out definitions"

def CR_start_find_invoice_page(f_CR_number):
	locate_move_and_click("procedures.png",0,0)
	time.sleep(1)
	locate_move_and_click('invoice_button.png',0,0)
	time.sleep(2)
	locate_move_and_click('add_page.png',57,50)
	time.sleep(2)
	locate_move_and_click('enter_position.png',0,20)
	pyautogui.press('backspace')
	pyautogui.press(['backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace','backspace'])
	pyautogui.press('down')
	pyautogui.press('down')
	pyautogui.press('space')
	pyautogui.typewrite(f_CR_number, interval=0.2)
	pyautogui.press('enter')
	pyautogui.hotkey('alt', 'y')
	time.sleep(3)
	locate_move_and_click('invoice_choose.png',0,0)
	locate_move_and_click("yes.png",0,0)
	time.sleep(3)

def check_bank_account():
	while True:
		account_choose = pyautogui.confirm(text='亲，是这个账号吗？', buttons=['YES', 'NO, NEXT'])
		if account_choose == 'YES':
			pyautogui.press('enter')
			break
		else:
			pyautogui.press('down')


def fullfil_invoice(f_seller_invoice_number,f_invoice_date,f_bank_account,f_total_amount,f_payment_type):
	"fullfilling invoice page"
	pyautogui.screenshot()
	(x,y)=pyautogui.center(pyautogui.locateOnScreen("original_invoice_number_position.png",confidence=0.8))
	pyautogui.moveTo(x, y, 0.5, pyautogui.easeOutQuad)
	pyautogui.move(100, 0)
	pyautogui.doubleClick()
	pyautogui.typewrite(f_seller_invoice_number, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_invoice_date, interval=0.15)
	pyautogui.press('enter')
	pyautogui.typewrite(f_invoice_date, interval=0.15)
	pyautogui.press('enter')
	pyautogui.press('enter')
	time.sleep(1)
	if bank_account != "0":
		pyautogui.typewrite(f_bank_account,interval=0.15)
		pyautogui.press('enter')
	else:
		locate_move_and_click("bank_account_position.png",215,0)
		time.sleep(2)
		pyautogui.press('down')
		check_bank_account()
		pyautogui.press('enter')
		time.sleep(1)
	locate_move_and_click("payment_type.png",90,0)
	pyautogui.typewrite(payment_type_dict[f_payment_type], interval=0.15)
	time.sleep(1)
	pyautogui.press('enter')
	locate_move_and_click("bar_position.png",0,0)
	pyautogui.dragRel(760, 0, duration=1)
	locate_move_and_click("change_number.png",0,0)
	time.sleep(2)
	pyautogui.move(0, 100)
	locate_move_and_click("total_amount.png",0,20)
	pyautogui.doubleClick()
	pyautogui.typewrite(f_total_amount, interval=0.15)
	pyautogui.hotkey('ctrl', 's')
	time.sleep(2)

def go_to_print_page():
	pyautogui.hotkey('ctrl', 'w')
	time.sleep(2)
	locate_move_and_click('yes.png',0,0)
	time.sleep(3)

def check_output_information():
	time.sleep(2)
	check_output = pyautogui.confirm(text='请确认请款单信息是否正确', buttons=['确认没毛病', '有点小问题'])
	if check_output == '确认没毛病':
		pass
	else:
		exit()

def check_default_delete():
	check_xls_delete = pyautogui.confirm(text='请确认桌面default.xls已删除', buttons=['已删除', '删除后再回来'])
	if check_xls_delete == '已删除':
		pass
	else:
		exit()

def first_time_print_excel():
	"print excel output"
	locate_move_and_click("excel_out.png",0,0)
	time.sleep(1)
	pyautogui.press('tab')
	pyautogui.press('enter')
	time.sleep(2)
	locate_move_and_click("desktop.png",0,0)
	time.sleep(1)
	locate_move_and_click("save_invoice_button.png",0,0)
	time.sleep(1)
	locate_move_and_click("yes.png",0,0)
	time.sleep(3)
	locate_move_and_click('yes.png',0,0)


def first_time_print_pdf_and_rename(f_quantity_tons):
	"print pdf output and rename"
	locate_move_and_click("pdf_out.png",0,0)
	time.sleep(1)
	pyautogui.press('tab')
	pyautogui.press('enter')
	time.sleep(2)
	locate_move_and_click("desktop.png",0,0)
	time.sleep(1)
	locate_move_and_click("save_invoice_button.png",0,0)
	time.sleep(1)
	pyautogui.press('tab')
	pyautogui.press('enter')
	time.sleep(3)
	locate_move_and_click('yes.png',0,0)
	new_name = '请款单' + str(1) + '-' + f_quantity_tons + '吨'
	desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
	os.rename(desktop_path + "/default.pdf",desktop_path + "/"+ new_name +".pdf")


def iterate_print_excel():
	"iterate to print excel output"
	locate_move_and_click("excel_out.png",0,0)
	time.sleep(2)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)
	pyautogui.hotkey('alt','o')
	time.sleep(2)
	locate_move_and_click("yes.png",0,0)
	time.sleep(2)

def iterate_print_pdf_and_rename(i,f_quantity_tons):
	"print pdf output and rename"
	locate_move_and_click("pdf_out.png",0,0)
	pyautogui.press('tab')
	pyautogui.press('tab')
	pyautogui.press('enter')
	time.sleep(4)
	locate_move_and_click("yes.png",0,0)
	new_name = '请款单' + str(i) + "-" + f_quantity_tons + '吨'
	desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
	os.rename(desktop_path + "/default.pdf",desktop_path + "/"+ new_name +".pdf")


def fulfil_and_print_all_invoices():
	if column_number == 3:
		check_default_delete()
		CR_start_find_invoice_page(CR_number)
		fullfil_invoice(seller_invoice_number,invoice_date,bank_account,total_amount,payment_type)
		time.sleep(2)
		go_to_print_page()
		check_output_information()
		time.sleep(3)
		first_time_print_excel()
		first_time_print_pdf_and_rename(quantity_tons)
		"close print page"
		locate_move_and_click('close_print_page.png',0,0)
		time.sleep(2)
	elif column_number > 3 and quantity_tons != "0":
		try:
			locate_move_and_click('close_invoice.png',36,0)
		except:
			pass
		CR_start_find_invoice_page(CR_number)
		fullfil_invoice(seller_invoice_number,invoice_date,bank_account,total_amount,payment_type)
		time.sleep(2)
		go_to_print_page()
		check_output_information()
		time.sleep(3)
		iterate_print_excel()
		iterate_print_pdf_and_rename(i,quantity_tons)
		"close print page"
		locate_move_and_click('close_print_page.png',0,0)
		time.sleep(2)

	else:
		pyautogui.alert("亲，都搞定了哦，建议保存CF号码")
		exit()


	# if payment_type != '冲估价采购':
	# 	iterate_print_excel()
	# 	iterate_print_pdf_and_rename(i,quantity_tons)
	# else:
	# 	pass
	# "go gack to procedure page"
	# time.sleep(2)
	# "close print page"
	# locate_move_and_click('close_print_page.png',0,0)

English_check()
start_point = pyautogui.confirm(text='请选择录入起点', buttons=[' CD环节 ', ' DH环节 ',' CF环节 '])
if start_point == ' CD环节 ':
	our_contract_number= str(NCsheet.cell(row=2,column=3).value)
	dealdate = str(NCsheet.cell(row=4,column=3).value)
	their_contract_number = str(NCsheet.cell(row=5,column=3).value)
	paymentrange = str(NCsheet.cell(row=6,column=3).value)
	dhdate = str(NCsheet.cell(row=9,column=3).value)
	barge_name = str(NCsheet.cell(row=10,column=3).value)
	tank_name= str(NCsheet.cell(row=11,column=3).value)
	quantity_tons = str(NCsheet.cell(row=12,column=3).value)
	quantity_bls = str(NCsheet.cell(row=13,column=3).value)
	bldate = str(NCsheet.cell(row=15,column=3).value)
	paymentdate = str(NCsheet.cell(row=16,column=3).value)
	QCnumber = str(NCsheet.cell(row=1,column=3).value)
	first_lot_all_the_way()
	first_iterate_from_DH_to_CR()

elif start_point == ' DH环节 ':
	CDnumber = 	str(NCsheet.cell(row=3,column=3).value)
	start_number = int(pyautogui.confirm(text='请在台账里录入CD号后选择开始批次', buttons=['2','3','4','5','6','7','8','9','10']))
	
	for i in range(start_number,NCsheet.max_column+1):
		column_number = i + 2
		dhdate = str(NCsheet.cell(row=9,column=column_number).value)
		barge_name = str(NCsheet.cell(row=10,column=column_number).value)
		tank_name= str(NCsheet.cell(row=11,column=column_number).value)
		quantity_tons = str(NCsheet.cell(row=12,column=column_number).value)
		quantity_bls = str(NCsheet.cell(row=13,column=column_number).value)
		bldate = str(NCsheet.cell(row=15,column=column_number).value)
		paymentdate = str(NCsheet.cell(row=16,column=column_number).value)
		seller_invoice_number= str(NCsheet.cell(row=18,column=column_number).value)
		total_amount = str(NCsheet.cell(row=19,column=column_number).value)
		payment_date = str(NCsheet.cell(row=20,column=column_number).value)
		invoice_date = str(NCsheet.cell(row=21,column=column_number).value)
		payment_type = str(NCsheet.cell(row=22,column=column_number).value)

		new_iterate_from_DH_to_CR()

else:
	start_number = int(pyautogui.confirm(text='请在台账里录入所有CR号后选择开始序号', buttons=['1','2','3','4','5','6','7','8','9','10']))
	
	for i in range(start_number,NCsheet.max_column+1):
		column_number = i + 2
		quantity_tons = str(NCsheet.cell(row=12,column=column_number).value)
		seller_invoice_number= str(NCsheet.cell(row=18,column=column_number).value)
		total_amount = str(NCsheet.cell(row=19,column=column_number).value)
		payment_date = str(NCsheet.cell(row=20,column=column_number).value)
		invoice_date = str(NCsheet.cell(row=21,column=column_number).value)
		payment_type = str(NCsheet.cell(row=22,column=column_number).value)
		CR_number = str(NCsheet.cell(row=14,column=column_number).value)
		bank_account = str(NCsheet.cell(row=17,column=column_number).value)

		fulfil_and_print_all_invoices()

"End of the program"

